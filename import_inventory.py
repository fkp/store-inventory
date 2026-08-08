#!/usr/bin/env python3
"""
Scout Inventory — Firebase Importer
====================================
Reads inventory_template.xlsx and upserts every row into Firestore.
Uses itemId as the document ID so re-running updates existing items
rather than creating duplicates.

SETUP
-----
1.  pip install firebase-admin openpyxl
2.  Download your service account key:
      Firebase console → Project Settings → Service accounts
      → Generate new private key → save as serviceAccountKey.json
      in the same folder as this script
3.  Put inventory_template.xlsx in the same folder
4.  Run:  python import_inventory.py

OPTIONS
-------
  --dry-run     Print what would be imported without writing to Firestore
  --clear       Delete ALL existing items before importing (fresh start)
  --file PATH   Use a different Excel file (default: inventory_template.xlsx)
"""

import argparse
import sys
import os

# ── Emoji mapping by category (auto-assigned)
CATEGORY_EMOJIS = {
    "Tents":      "⛺",
    "Camping":    "🏕️",
    "Cooking":    "🍳",
    "Games":      "🎮",
    "Stationery": "✏️",
    "First Aid":  "🩺",
    "Navigation": "🧭",
    "Safety":     "🦺",
    "Transport":  "🎒",
    "Electronic": "📡",
    "Admin":      "📋",
    "Misc":       "📦",
    "Other":      "📦",
}

VALID_CATEGORIES = set(CATEGORY_EMOJIS.keys())

def get_emoji(category):
    return CATEGORY_EMOJIS.get(category, "📦")

def read_excel(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath, data_only=True)
    ws = wb["Inventory"]

    headers = [cell.value for cell in ws[1]]
    required = {"itemId", "name", "category", "total"}
    missing = required - set(headers)
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        sys.exit(1)

    col = {h: i for i, h in enumerate(headers)}
    items = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip entirely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        item_id  = row[col["itemId"]]
        name     = row[col["name"]]
        category = row[col["category"]]
        total    = row[col["total"]]
        room     = row[col.get("room", -1)] if "room" in col else ""
        location = row[col.get("location", -1)] if "location" in col else ""
        notes    = row[col.get("notes", -1)] if "notes" in col else ""

        # Validate
        row_errors = []
        if item_id is None:
            row_errors.append("itemId is empty")
        if not name:
            row_errors.append("name is empty")
        if not category:
            row_errors.append("category is empty")
        elif category not in VALID_CATEGORIES:
            row_errors.append(f"unknown category '{category}' — valid: {sorted(VALID_CATEGORIES)}")
        if total is None:
            row_errors.append("total is empty")
        elif not isinstance(total, (int, float)) or total < 0:
            row_errors.append(f"total must be a positive number, got: {total}")

        if row_errors:
            errors.append(f"  Row {row_num}: {'; '.join(row_errors)}")
            continue

        items.append({
            "itemId":   str(int(item_id)),
            "name":     str(name).strip(),
            "category": str(category).strip(),
            "total":    int(total),
            "room":     str(room).strip() if room else "",
            "location": str(location).strip() if location else "",
            "notes":    str(notes).strip() if notes else "",
            "emoji":    get_emoji(str(category).strip()),
        })

    return items, errors

def main():
    parser = argparse.ArgumentParser(description="Import inventory Excel → Firestore")
    parser.add_argument("--dry-run", action="store_true", help="Print items without writing to Firestore")
    parser.add_argument("--clear",   action="store_true", help="Delete all existing Firestore items before import")
    parser.add_argument("--file",    default="inventory_template.xlsx", help="Path to Excel file")
    parser.add_argument("--key",     default="serviceAccountKey.json",  help="Path to Firebase service account key")
    args = parser.parse_args()

    # ── Read Excel
    print(f"\n📂 Reading {args.file}…")
    items, errors = read_excel(args.file)

    if errors:
        print(f"\n⚠️  Validation errors found ({len(errors)}):")
        for e in errors:
            print(e)
        if not items:
            print("\nNo valid items to import. Fix the errors above and try again.")
            sys.exit(1)
        print(f"\nContinuing with {len(items)} valid item(s), skipping {len(errors)} error(s).\n")
    else:
        print(f"✓ {len(items)} items read, no errors.\n")

    # ── Dry run — just print
    if args.dry_run:
        print("DRY RUN — nothing will be written to Firestore.\n")
        print(f"{'ID':<6} {'Name':<40} {'Category':<12} {'Total':<6} {'Room':<20} {'Location':<20} Emoji")
        print("-" * 110)
        for item in items:
            print(f"{item['itemId']:<6} {item['name']:<40} {item['category']:<12} {item['total']:<6} {item['room']:<20} {item['location']:<20} {item['emoji']}")
        print(f"\nWould upsert {len(items)} items.")
        return

    # ── Connect to Firebase
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        print("ERROR: firebase-admin not installed. Run: pip install firebase-admin")
        sys.exit(1)

    if not os.path.exists(args.key):
        print(f"ERROR: Service account key not found: {args.key}")
        print("Download it from Firebase console → Project Settings → Service accounts → Generate new private key")
        sys.exit(1)

    print(f"🔑 Connecting to Firebase using {args.key}…")
    cred = credentials.Certificate(args.key)
    firebase_admin.initialize_app(cred)
    db = firestore.client()
    items_col = db.collection("items")

    # ── Optionally clear existing items
    if args.clear:
        print("🗑️  Clearing existing items from Firestore…")
        existing = items_col.stream()
        deleted = 0
        for doc in existing:
            doc.reference.delete()
            deleted += 1
        print(f"   Deleted {deleted} existing item(s).\n")

    # ── Upsert items
    print(f"⬆️  Upserting {len(items)} items…\n")
    success = 0
    failed  = 0

    for item in items:
        doc_id = f"item_{item['itemId']}"
        data = {
            "name":     item["name"],
            "category": item["category"],
            "total":    item["total"],
            "room":     item["room"],
            "location": item["location"],
            "notes":    item["notes"],
            "emoji":    item["emoji"],
            "itemId":   int(item["itemId"]),
        }
        try:
            items_col.document(doc_id).set(data)  # set() = upsert (overwrites)
            print(f"  ✓ [{doc_id}] {item['name']}")
            success += 1
        except Exception as e:
            print(f"  ✗ [{doc_id}] {item['name']} — ERROR: {e}")
            failed += 1

    # ── Summary
    print(f"\n{'='*50}")
    print(f"✓ {success} items upserted successfully")
    if failed:
        print(f"✗ {failed} items failed — check errors above")
    print(f"{'='*50}\n")
    print("Done! Open your inventory app to see the updated data.")

if __name__ == "__main__":
    main()
